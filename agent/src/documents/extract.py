"""Best-effort text extraction from uploaded finance documents."""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from xml.etree import ElementTree


def extract_text(raw: bytes, kind: str) -> tuple[str, float]:
    """Return (text, confidence) for a detected upload kind."""
    if not raw:
        return "", 0.0

    if kind in {"txt", "md", "csv"}:
        text = _decode_text(raw)
        if kind == "csv":
            return _csv_to_text(text), 0.95
        return text, 0.98

    if kind in {"json", "jsonl"}:
        text = _decode_text(raw)
        return text, 0.95

    if kind == "pdf":
        return _extract_pdf_text(raw), 0.75

    if kind == "docx":
        return _extract_docx_text(raw), 0.85

    if kind in {"xlsx", "xls"}:
        return _extract_spreadsheet_text(raw, kind), 0.9

    return _decode_text(raw), 0.5


def _decode_text(raw: bytes) -> str:
    return raw.decode("utf-8-sig", errors="replace")


def _csv_to_text(text: str) -> str:
    rows: list[str] = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        rows.append(" | ".join(cell.strip() for cell in row if cell is not None))
    return "\n".join(rows)


def _extract_pdf_text(raw: bytes) -> str:
    blob = raw.decode("latin-1", errors="ignore")
    parts = re.findall(r"\(([^()\\]*(?:\\.[^()\\]*)*)\)", blob)
    cleaned = []
    for part in parts:
        value = part.replace("\\n", "\n").replace("\\r", "").strip()
        if len(value) >= 2 and not value.startswith("/"):
            cleaned.append(value)
    return "\n".join(cleaned)


def _extract_docx_text(raw: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    texts = []
    for node in root.iter():
        if node.tag.endswith("}t") and node.text:
            texts.append(node.text)
    return "\n".join(texts)


def _extract_spreadsheet_text(raw: bytes, kind: str) -> str:
    if kind == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                if cells:
                    lines.append(" | ".join(cells))
        workbook.close()
        return "\n".join(lines)

    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    lines: list[str] = []
    for sheet in book.sheets():
        lines.append(f"# {sheet.name}")
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col_idx)).strip() for col_idx in range(sheet.ncols)]
            cells = [cell for cell in cells if cell]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def chunk_text(text: str, *, size: int = 800, overlap: int = 100) -> list[str]:
    normalized = re.sub(r"\n{3,}", "\n\n", text.strip())
    if not normalized:
        return []
    if len(normalized) <= size:
        return [normalized]
    chunks: list[str] = []
    start = 0
    while start < len(normalized):
        end = min(len(normalized), start + size)
        chunks.append(normalized[start:end].strip())
        if end >= len(normalized):
            break
        start = max(0, end - overlap)
    return [chunk for chunk in chunks if chunk]
