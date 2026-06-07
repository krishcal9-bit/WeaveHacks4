"""
Connector registry + structured file parsers for finance-operations feeds.

Each connector is configured by a single environment variable that points at a
CSV or JSON export. If the variable is unset the connector is **not configured**
(it never invents data); if it is set but the file is missing, that is surfaced
as an explicit blocker. Parsing is delegated to the typed models in
:mod:`src.integrations.models` via the stdlib ``csv``/``json`` parsers — there is
no ad hoc line-splitting.

Future API-backed connectors (Stripe, NetSuite, Salesforce, …) can be added by
extending :data:`CONNECTORS`; until one is genuinely wired and verified it must
keep ``Origin.LIVE_API`` out of its reported status.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

from pydantic import ValidationError

from src.integrations.models import (
    RECORD_MODELS,
    Origin,
    SourceFormat,
    SourceType,
    ValidationIssue,
    _SourceRecord,
)


@dataclass(frozen=True)
class ConnectorSpec:
    connector_id: str
    source_type: SourceType
    env_var: str
    fixture_filename: str
    description: str

    @property
    def record_model(self) -> type[_SourceRecord]:
        return RECORD_MODELS[self.source_type]


@dataclass(frozen=True)
class ParseMetadata:
    """Source parser details that should travel with persisted provenance."""

    workbook_name: Optional[str] = None
    workbook_sheet: Optional[str] = None
    workbook_sheets: tuple[str, ...] = ()
    header_row_number: Optional[int] = None
    hidden_column_count: int = 0
    extra_column_count: int = 0


# The connector taxonomy. connector_id == source_type value for a 1:1 mapping.
CONNECTORS: dict[str, ConnectorSpec] = {
    "ledger": ConnectorSpec(
        "ledger", SourceType.LEDGER, "ATLAS_LEDGER_FILE", "ledger.csv",
        "General ledger / cash transactions (CSV or JSON).",
    ),
    "invoices": ConnectorSpec(
        "invoices", SourceType.INVOICES, "ATLAS_INVOICES_FILE", "invoices.csv",
        "Accounts-payable vendor invoices (CSV or JSON).",
    ),
    "vendor_export": ConnectorSpec(
        "vendor_export", SourceType.VENDOR_EXPORT, "ATLAS_VENDOR_EXPORT_FILE", "vendor_export.json",
        "Vendor / procurement contract export (CSV or JSON).",
    ),
    "crm_opportunities": ConnectorSpec(
        "crm_opportunities", SourceType.CRM_OPPORTUNITIES, "ATLAS_CRM_FILE", "crm_opportunities.csv",
        "CRM pipeline opportunity export (CSV or JSON).",
    ),
    "headcount_plan": ConnectorSpec(
        "headcount_plan", SourceType.HEADCOUNT_PLAN, "ATLAS_HEADCOUNT_FILE", "headcount_plan.csv",
        "Actual / updated headcount plan from HRIS or planning sheet (CSV or JSON).",
    ),
    "security_evidence": ConnectorSpec(
        "security_evidence", SourceType.SECURITY_EVIDENCE, "ATLAS_SECURITY_FILE", "security_evidence.json",
        "Security / compliance control evidence (CSV or JSON).",
    ),
    "board_policy": ConnectorSpec(
        "board_policy", SourceType.BOARD_POLICY, "ATLAS_BOARD_POLICY_FILE", "board_policies.json",
        "Board policy / constraint documents with optional machine-checkable rules (CSV or JSON).",
    ),
}


EXCEL_FORMATS = {SourceFormat.XLSX, SourceFormat.XLS}

SHEET_ALIASES: dict[SourceType, tuple[str, ...]] = {
    SourceType.LEDGER: ("ledger", "general ledger", "gl", "cash", "bank", "transactions", "cloudledger"),
    SourceType.INVOICES: ("invoice", "invoices", "ap", "a p", "payable", "payables", "bills", "payablesdesk"),
    SourceType.VENDOR_EXPORT: ("vendor", "vendors", "supplier", "suppliers", "contract", "contracts", "procurement", "contractvault"),
    SourceType.CRM_OPPORTUNITIES: ("crm", "opportunity", "opportunities", "pipeline", "forecast", "sales", "pipelinehub"),
    SourceType.HEADCOUNT_PLAN: ("headcount", "hiring", "workforce", "people", "hris", "peopleroster"),
    SourceType.SECURITY_EVIDENCE: ("security", "evidence", "soc", "soc2", "grc", "controls", "trustvault", "risk"),
    SourceType.BOARD_POLICY: ("board policy", "board policies", "policy", "policies", "rules", "governance", "boardportal"),
}

HEADER_ALIASES: dict[SourceType, dict[str, tuple[str, ...]]] = {
    SourceType.LEDGER: {
        "txn_id": ("txn_id", "transaction_id", "transaction id", "id", "reference", "ref_number", "bank_reference"),
        "date": ("date", "posted_date", "posting_date", "transaction_date", "settled_date"),
        "account": ("account", "account_name", "bank_account", "ledger_account"),
        "description": ("description", "bank_description", "statement_description", "memo", "name"),
        "amount": ("amount", "signed_amount", "net_amount"),
        "category": ("category", "source_category", "bank_category", "gl_category"),
        "vendor_id": ("vendor_id", "vendor id", "supplier_id"),
        "vendor_name": ("vendor_name", "vendor", "payee", "merchant_name", "counterparty"),
    },
    SourceType.INVOICES: {
        "invoice_id": ("invoice_id", "invoice id", "invoice #", "invoice number", "bill id"),
        "vendor_id": ("vendor_id", "vendor id", "supplier_id"),
        "vendor_name": ("vendor_name", "vendor", "supplier", "payee"),
        "issue_date": ("issue_date", "issue date", "invoice date", "bill date"),
        "due_date": ("due_date", "due date", "payment due"),
        "amount": ("amount", "invoice amount", "total", "gross amount"),
        "po_number": ("po_number", "po number", "purchase order", "po"),
    },
    SourceType.VENDOR_EXPORT: {
        "vendor_id": ("vendor_id", "vendor id", "supplier_id", "contract id"),
        "name": ("name", "vendor name", "supplier", "contract name"),
        "annual_cost": ("annual_cost", "annual cost", "annual value", "contract value", "arr"),
        "monthly_cost": ("monthly_cost", "monthly cost", "monthly value"),
        "renewal_date": ("renewal_date", "renewal date", "end date", "contract end"),
    },
    SourceType.CRM_OPPORTUNITIES: {
        "opportunity_id": ("opportunity_id", "opportunity id", "opp id", "deal id"),
        "name": ("name", "opportunity name", "deal name"),
        "stage": ("stage", "sales stage"),
        "arr": ("arr", "annual recurring revenue", "amount", "deal arr"),
        "probability": ("probability", "probability %", "probability_pct", "probability pct"),
        "weighted_arr": ("weighted_arr", "weighted arr", "forecast arr"),
        "close_date": ("close_date", "close date", "expected close"),
    },
    SourceType.HEADCOUNT_PLAN: {
        "role_id": ("role_id", "role id", "position id", "req id"),
        "team": ("team", "department", "dept", "org"),
        "role": ("role", "title", "job title"),
        "headcount": ("headcount", "hc", "count", "openings"),
        "monthly_cost": ("monthly_cost", "monthly cost", "salary monthly"),
        "start_month": ("start_month", "start month"),
        "current_start_date": ("current_start_date", "current start date", "start date"),
    },
    SourceType.SECURITY_EVIDENCE: {
        "control_id": ("control_id", "control id", "control", "control ref"),
        "framework": ("framework", "standard"),
        "title": ("title", "control title", "requirement"),
        "status": ("status", "control status", "evidence status"),
        "evidence_date": ("evidence_date", "evidence date", "last evidence"),
    },
    SourceType.BOARD_POLICY: {
        "policy_id": ("policy_id", "policy id", "policy", "rule id"),
        "title": ("title", "policy title", "rule title"),
        "category": ("category", "policy category"),
        "text": ("text", "policy text", "description", "policy"),
        "rule": ("rule", "machine rule", "rule key"),
    },
}


def fixtures_dir() -> Path:
    """Directory holding the opt-in Acme demo operating-data fixtures."""
    return Path(__file__).resolve().parents[1] / "data" / "fixtures"


def configured_path(spec: ConnectorSpec) -> Optional[str]:
    """The source path from the connector's env var, or None when unset."""
    value = os.getenv(spec.env_var, "").strip()
    return value or None


def fixture_path(spec: ConnectorSpec) -> Path:
    return fixtures_dir() / spec.fixture_filename


def resolve_origin(path: Path) -> Origin:
    """Classify a file path as a bundled demo fixture vs. an external export."""
    try:
        path.resolve().relative_to(fixtures_dir().resolve())
        return Origin.ACME_DEMO_FIXTURE
    except (ValueError, OSError):
        return Origin.EXTERNAL_FILE


def detect_format(path: Path, override: Optional[SourceFormat] = None) -> SourceFormat:
    if override is not None:
        return override
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return SourceFormat.CSV
    if suffix in (".json", ".jsonl"):
        return SourceFormat.JSON
    if suffix == ".xlsx":
        return SourceFormat.XLSX
    if suffix == ".xls":
        return SourceFormat.XLS
    raise ValueError(f"unsupported source format for {path.name!r}; use .csv, .json, .xlsx, or .xls")


def checksum_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def file_timestamp(path: Path) -> datetime:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)


def _rows_from_csv(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _rows_from_json(raw: bytes, source_type: SourceType) -> list[dict[str, Any]]:
    data = json.loads(raw.decode("utf-8-sig"))
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        # Accept {"records": [...]} or {"<source_type>": [...]} envelopes.
        rows = data.get("records")
        if rows is None:
            rows = data.get(source_type.value)
        if rows is None:
            raise ValueError(
                f"JSON object must contain a 'records' or '{source_type.value}' array"
            )
    else:
        raise ValueError("JSON source must be an array or an object with a 'records' array")
    if not isinstance(rows, list):
        raise ValueError("expected a JSON array of records")
    return [row for row in rows if isinstance(row, dict)]


def _norm_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _sheet_name_score(source_type: SourceType, sheet_name: str) -> int:
    normalized = re.sub(r"[^a-z0-9]+", " ", sheet_name.lower()).strip()
    compact = _norm_label(sheet_name)
    score = 0
    for alias in SHEET_ALIASES.get(source_type, ()):
        alias_norm = re.sub(r"[^a-z0-9]+", " ", alias.lower()).strip()
        alias_compact = _norm_label(alias)
        if compact == alias_compact:
            score = max(score, 80)
        elif alias_norm and alias_norm in normalized:
            score = max(score, 60)
    return score


def _header_map(spec: ConnectorSpec) -> dict[str, str]:
    out: dict[str, str] = {}
    for field in spec.record_model.model_fields:
        out[_norm_label(field)] = field
        out[_norm_label(field.replace("_", " "))] = field
    for canonical, aliases in HEADER_ALIASES.get(spec.source_type, {}).items():
        out[_norm_label(canonical)] = canonical
        for alias in aliases:
            out[_norm_label(alias)] = canonical
    return out


def _canonicalize_header(spec: ConnectorSpec, value: Any, index: int) -> tuple[str, bool]:
    text = str(value or "").strip()
    if not text:
        return f"column_{index + 1}", False
    canonical = _header_map(spec).get(_norm_label(text))
    if canonical:
        return canonical, True
    snake = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return snake or f"column_{index + 1}", False


def _cell_value(value: Any) -> Any:
    if value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_is_blank(row: list[Any]) -> bool:
    return all(value in (None, "") for value in row)


def _detect_header(rows: list[list[Any]], spec: ConnectorSpec) -> tuple[int, list[str], int, int]:
    fields = set(spec.record_model.model_fields)
    best: tuple[int, int, list[str], int, int] | None = None
    for index, row in enumerate(rows[: min(len(rows), 20)]):
        canonical_headers: list[str] = []
        recognized = 0
        extra = 0
        seen_headers: set[str] = set()
        for col_index, raw_header in enumerate(row):
            header, is_known = _canonicalize_header(spec, raw_header, col_index)
            if header in seen_headers:
                header = f"{header}_{col_index + 1}"
                is_known = False
            seen_headers.add(header)
            canonical_headers.append(header)
            if is_known and header in fields:
                recognized += 1
            elif raw_header not in (None, ""):
                extra += 1
        if recognized < 2:
            continue
        # Prefer rows with many known headers; earlier rows break ties.
        score = recognized * 10 - extra
        candidate = (score, index, canonical_headers, recognized, extra)
        if best is None or candidate[0] > best[0]:
            best = candidate
    if best is None:
        raise ValueError(f"Could not detect a header row for {spec.source_type.value}; expected columns such as {', '.join(sorted(list(fields))[:6])}.")
    _, index, headers, recognized, extra = best
    return index, headers, recognized, extra


def _rows_from_matrix(
    rows: list[list[Any]],
    spec: ConnectorSpec,
    *,
    sheet_name: str,
    workbook_sheets: tuple[str, ...],
    hidden_column_count: int,
) -> tuple[list[dict[str, Any]], ParseMetadata, int]:
    header_index, headers, recognized, extra_columns = _detect_header(rows, spec)
    out: list[dict[str, Any]] = []
    for raw_row in rows[header_index + 1 :]:
        if _row_is_blank(raw_row):
            continue
        padded = raw_row + [None] * max(0, len(headers) - len(raw_row))
        row = {header: _cell_value(padded[index]) for index, header in enumerate(headers)}
        out.append(row)
    metadata = ParseMetadata(
        workbook_sheet=sheet_name,
        workbook_sheets=workbook_sheets,
        header_row_number=header_index + 1,
        hidden_column_count=hidden_column_count,
        extra_column_count=extra_columns,
    )
    return out, metadata, recognized


def _load_xlsx_sheets(raw: bytes) -> list[tuple[str, list[list[Any]], int]]:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    sheets: list[tuple[str, list[list[Any]], int]] = []
    for ws in wb.worksheets:
        hidden_columns = 0
        max_column = ws.max_column or 0
        for index in range(1, max_column + 1):
            if ws.column_dimensions[get_column_letter(index)].hidden:
                hidden_columns += 1
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        sheets.append((ws.title, rows, hidden_columns))
    return sheets


def _load_xls_sheets(raw: bytes) -> list[tuple[str, list[list[Any]], int]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=raw, formatting_info=True)
    sheets: list[tuple[str, list[list[Any]], int]] = []
    for sheet in book.sheets():
        hidden_columns = sum(1 for info in sheet.colinfo_map.values() if getattr(info, "hidden", 0))
        rows: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            values: list[Any] = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode).date().isoformat()
                    except Exception:
                        pass
                values.append(value)
            rows.append(values)
        sheets.append((sheet.name, rows, hidden_columns))
    return sheets


def _rows_from_excel(raw: bytes, fmt: SourceFormat, spec: ConnectorSpec) -> tuple[list[dict[str, Any]], ParseMetadata]:
    sheets = _load_xlsx_sheets(raw) if fmt is SourceFormat.XLSX else _load_xls_sheets(raw)
    workbook_sheets = tuple(name for name, _, _ in sheets)
    candidates: list[tuple[int, int, str, list[dict[str, Any]], ParseMetadata]] = []
    failures: list[str] = []
    for order, (sheet_name, matrix, hidden_columns) in enumerate(sheets):
        if not matrix or all(_row_is_blank(row) for row in matrix):
            continue
        try:
            rows, metadata, recognized_headers = _rows_from_matrix(
                matrix,
                spec,
                sheet_name=sheet_name,
                workbook_sheets=workbook_sheets,
                hidden_column_count=hidden_columns,
            )
        except ValueError as exc:
            failures.append(f"{sheet_name}: {exc}")
            continue
        sheet_score = _sheet_name_score(spec.source_type, sheet_name)
        if sheet_score <= 0 and recognized_headers < 3:
            continue
        score = sheet_score + recognized_headers * 12 - metadata.extra_column_count
        candidates.append((score, -order, sheet_name, rows, metadata))
    if not candidates:
        hint = "; ".join(failures[:3]) if failures else f"workbook sheets: {', '.join(workbook_sheets) or 'none'}"
        raise ValueError(f"No worksheet matched connector {spec.connector_id}. Rename a sheet for this connector or include recognizable headers. {hint}")
    candidates.sort(reverse=True)
    _, _, _, rows, metadata = candidates[0]
    return rows, metadata


def parse_records(
    spec: ConnectorSpec,
    raw: bytes,
    fmt: SourceFormat,
) -> tuple[list[_SourceRecord], list[ValidationIssue], int]:
    records, issues, duplicates, _ = parse_records_with_metadata(spec, raw, fmt)
    return records, issues, duplicates


def parse_records_with_metadata(
    spec: ConnectorSpec,
    raw: bytes,
    fmt: SourceFormat,
) -> tuple[list[_SourceRecord], list[ValidationIssue], int, ParseMetadata]:
    """Parse + validate raw bytes into typed records.

    Returns ``(records, issues, duplicate_count)``. Rows that fail validation are
    reported as :class:`ValidationIssue` (with row/field context) and dropped;
    duplicate ``record_key`` values keep the first occurrence and are counted.
    """
    metadata = ParseMetadata()
    if fmt is SourceFormat.CSV:
        rows = _rows_from_csv(raw)
    elif fmt is SourceFormat.JSON:
        rows = _rows_from_json(raw, spec.source_type)
    elif fmt in EXCEL_FORMATS:
        rows, metadata = _rows_from_excel(raw, fmt, spec)
    else:
        raise ValueError(f"unsupported source format: {fmt}")

    model = spec.record_model
    records: list[_SourceRecord] = []
    issues: list[ValidationIssue] = []
    seen: set[str] = set()
    duplicates = 0

    for index, row in enumerate(rows):
        if fmt is SourceFormat.CSV:
            location = f"row {index + 2}"
        elif fmt in EXCEL_FORMATS and metadata.header_row_number is not None:
            location = f"{metadata.workbook_sheet or 'worksheet'} row {metadata.header_row_number + index + 1}"
        else:
            location = f"record {index + 1}"
        try:
            record = model.model_validate(row)
        except ValidationError as exc:
            for err in exc.errors():
                field = ".".join(str(part) for part in err.get("loc", ())) or None
                issues.append(ValidationIssue(location=location, field=field, message=err.get("msg", "invalid")))
            continue
        except (ValueError, TypeError) as exc:
            issues.append(ValidationIssue(location=location, message=str(exc)))
            continue
        key = record.record_key()
        if key in seen:
            duplicates += 1
            if spec.source_type is not SourceType.INVOICES:
                issues.append(
                    ValidationIssue(
                        location=location,
                        message=f"duplicate record key {key!r}; kept the first occurrence",
                    )
                )
                continue
            # Duplicate invoice ids are material finance evidence, not just an
            # import concern. Keep the row so reconciliation can flag the AP risk.
        else:
            seen.add(key)
        records.append(record)

    return records, issues, duplicates, metadata
