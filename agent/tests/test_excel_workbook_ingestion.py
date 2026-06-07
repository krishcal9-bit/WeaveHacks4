from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import Workbook

from src import redis_layer as R
from src.api import router
from src.data.seed import seed
from src.integrations import connectors as C
from src.integrations import service as OPS
from src.integrations.models import SourceFormat


def _xlsx_bytes(sheets: dict[str, list[list[Any]]], hidden_columns: dict[str, list[str]] | None = None) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        for row in rows:
            ws.append(row)
        for column in (hidden_columns or {}).get(title, []):
            ws.column_dimensions[column].hidden = True
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ops_workbook() -> bytes:
    return _xlsx_bytes(
        {
            "Read Me": [["Atlas operations workbook"], ["Generated", "2026-06-07"]],
            "CloudLedger GL": [
                ["Export", "CloudLedger"],
                [],
                ["Reference", "Transaction Date", "Account Name", "Description", "Statement Description", "Debit", "Credit", "Payee", "Internal Notes"],
                ["TXN-1", "2026-06-03", "Corporate Card", "Datadog card charge", "CARD 4242 DATADOG INC NY", "21500", None, "Data Dog Inc", "hidden note"],
                ["TXN-2", "2026-06-04", "Corporate Card", "Unknown SaaS", "CARD 4242 UNKNOWN_VENDOR TOOL", "2200", None, "Unknown Vendor SaaS Tool", "needs coding"],
            ],
            "AP Aging - Invoices": [
                ["Source", "PayablesDesk"],
                ["Invoice ID", "Vendor ID", "Vendor Name", "Issue Date", "Due Date", "Amount", "Currency", "Status", "Terms", "Source System", "Extra Approval Note"],
                ["INV-X1", "datadog", "Datadog", "2026-06-01", "2026-06-06", "21500", "USD", "open", "Due on receipt", "NetSuite", "manager says ok"],
                ["INV-X2", "acme-analytics", "Acme Analytics", "2026-06-02", None, "120000", "USD", "open", "Net 15", "Procurement inbox", "missing due date"],
            ],
            "ContractVault Vendors": [
                ["Vendor ID", "Vendor Name", "Annual Cost", "Monthly Cost", "Renewal Date", "Status", "Owner", "Billing Frequency", "Board Approved", "Security Clause"],
                ["datadog", "Data Dog, Inc.", 180000, 15000, "2026-07-25", "up_for_renewal", "Platform Ops", "annual", True, "SOC 2 current"],
                ["acme-analytics", "Acme Analytics Platform", 250000, 20833, "2027-06-01", "active", "Data Ops", "annual", False, "DPA unsigned"],
            ],
            "PipelineHub CRM": [
                ["Opportunity ID", "Opportunity Name", "Account", "Stage", "ARR", "Probability", "Weighted ARR", "Close Date", "Owner"],
                ["OPP-X1", "Beacon Pilot", "Beacon Logistics", "Technical validation", 760000, "40%", 304000, "2026-11-20", "J. Chen"],
            ],
            "PeopleRoster Headcount": [
                ["Role ID", "Team", "Role", "Headcount", "Monthly Cost", "Current Start Date", "Status", "Approval Status"],
                ["HC-X1", "Eng", "Platform Engineer", 1, 21000, "2026-09-15", "open", "partial"],
            ],
            "TrustVault Security": [
                ["Control ID", "Framework", "Title", "Status", "Evidence Date", "Blocks Revenue", "Blocked ARR", "Summary"],
                ["CC7.2", "SOC 2", "Change management evidence", "gap", "2026-01-15", True, 310000, "Evidence stale"],
            ],
            "BoardPortal Policies": [
                ["Policy ID", "Title", "Category", "Policy Text", "Rule", "Threshold", "Severity"],
                ["BP-X", "Board approval for new vendor commitments", "vendor", "New vendor commitments above 100000 require board approval.", "vendor_commitment_board_notification", 100000, "high"],
            ],
        },
        hidden_columns={"CloudLedger GL": ["I"]},
    )


def test_xlsx_parser_detects_sheet_header_hidden_and_extra_columns() -> None:
    raw = _xlsx_bytes(
        {
            "Overview": [["Not data"]],
            "AP Aging - Invoices": [
                ["Generated", "PayablesDesk"],
                [],
                ["Invoice ID", "Vendor Name", "Issue Date", "Due Date", "Amount", "Currency", "Status", "Hidden Internal Notes"],
                ["INV-1", "Datadog", "2026-06-01", "2026-06-06", "21500", "USD", "open", "not imported as a model field"],
            ],
        },
        hidden_columns={"AP Aging - Invoices": ["H"]},
    )

    records, issues, duplicates, metadata = C.parse_records_with_metadata(
        C.CONNECTORS["invoices"],
        raw,
        SourceFormat.XLSX,
    )

    assert len(records) == 1
    assert issues == []
    assert duplicates == 0
    assert metadata.workbook_sheet == "AP Aging - Invoices"
    assert metadata.header_row_number == 3
    assert metadata.hidden_column_count == 1
    assert metadata.extra_column_count == 2
    dumped = records[0].model_dump(mode="json")
    assert dumped["invoice_id"] == "INV-1"
    assert dumped["amount"] == 21500.0
    assert "hidden_internal_notes" not in dumped


def test_xls_parser_supports_legacy_workbooks_with_header_detection() -> None:
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet("CloudLedger GL")
    rows = [
        ["Legacy Excel export"],
        [],
        ["Reference", "Transaction Date", "Account Name", "Description", "Statement Description", "Debit", "Payee", "Hidden Comment"],
        ["TXN-XLS", "2026-06-03", "Corporate Card", "Datadog charge", "CARD 4242 DATADOG INC NY", "21500", "Data Dog Inc", "legacy hidden"],
    ]
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            ws.write(row_index, col_index, value)
    ws.col(7).hidden = True
    buf = BytesIO()
    wb.save(buf)

    records, issues, duplicates, metadata = C.parse_records_with_metadata(
        C.CONNECTORS["ledger"],
        buf.getvalue(),
        SourceFormat.XLS,
    )

    assert issues == []
    assert duplicates == 0
    assert len(records) == 1
    assert metadata.workbook_sheet == "CloudLedger GL"
    assert metadata.header_row_number == 3
    assert metadata.hidden_column_count == 1
    assert metadata.extra_column_count == 1
    row = records[0].model_dump(mode="json")
    assert row["txn_id"] == "TXN-XLS"
    assert row["inferred_vendor_id"] == "datadog"
    assert row["normalized_category"] == "software"


def test_single_workbook_import_persists_all_connector_provenance_and_council_evidence() -> None:
    assert R.ping(), "Redis must be running for workbook ingestion coverage"
    seed(verbose=False, include_company=True)
    OPS.reset_demo_state()

    results = OPS.import_workbook(source_name="atlas-ops-workbook.xlsx", raw=_ops_workbook())
    report = OPS.run_reconciliation()

    assert len(results) == 7
    assert all(result.provenance.has_data() for result in results)
    by_connector = {result.provenance.connector_id: result.provenance for result in results}
    assert by_connector["ledger"].source_format.value == "xlsx"
    assert by_connector["ledger"].workbook_name == "atlas-ops-workbook.xlsx"
    assert by_connector["ledger"].workbook_sheet == "CloudLedger GL"
    assert by_connector["ledger"].header_row_number == 3
    assert by_connector["ledger"].hidden_column_count == 1
    assert by_connector["ledger"].extra_column_count == 1
    assert by_connector["invoices"].workbook_sheet == "AP Aging - Invoices"
    assert set(by_connector["invoices"].workbook_sheets) >= {"CloudLedger GL", "AP Aging - Invoices", "BoardPortal Policies"}

    source = OPS.get_source("ledger")
    assert source is not None
    assert source["provenance"]["workbook_sheet"] == "CloudLedger GL"
    assert source["sample"][0]["raw_description"] == "Datadog card charge"
    assert "internal_notes" not in source["sample"][0]

    snapshot = OPS.operations_context_snapshot()
    assert snapshot is not None
    ledger_evidence = next(item for item in snapshot["sources"] if item["source_type"] == "ledger")
    assert ledger_evidence["workbook_name"] == "atlas-ops-workbook.xlsx"
    assert ledger_evidence["workbook_sheet"] == "CloudLedger GL"
    assert ledger_evidence["header_row_number"] == 3
    assert report.confidence.sources_imported == report.confidence.sources_total


def test_workbook_upload_api_validation_and_response_shape() -> None:
    assert R.ping(), "Redis must be running for workbook upload API coverage"
    seed(verbose=False, include_company=True)
    OPS.reset_demo_state()
    client = TestClient(router)

    bad = client.post("/api/connectors/import-workbook", files={"file": ("sources.csv", b"a,b\n1,2\n", "text/csv")})
    assert bad.status_code == 400
    assert ".xlsx" in bad.json()["detail"]

    response = client.post(
        "/api/connectors/import-workbook",
        files={
            "file": (
                "atlas-ops-workbook.xlsx",
                _ops_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert set(payload["workbook"]["imported_connectors"]) == set(C.CONNECTORS)
    assert payload["workbook"]["failed_connectors"] == []
    ledger = next(connector for connector in payload["connectors"] if connector["connector_id"] == "ledger")
    assert ledger["workbook_name"] == "atlas-ops-workbook.xlsx"
    assert ledger["workbook_sheet"] == "CloudLedger GL"
    assert ledger["source_format"] == "xlsx"
